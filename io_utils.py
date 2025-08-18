import csv
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

import os
import pandas as pd
from openpyxl import Workbook, load_workbook

# Configuration for file handling
DEFAULT_PROMPT_FILENAME = os.getenv("DEFAULT_PROMPT_FILENAME", "prompt.txt")
ALT_PROMPT_SUFFIX = os.getenv("ALT_PROMPT_SUFFIX", ".prompt.txt")
OUTPUT_FILE_SUFFIX = os.getenv("OUTPUT_FILE_SUFFIX", "_enriched.csv")

# ---------------------------------------------------------------------------
# Generic table loaders/writers
# ---------------------------------------------------------------------------

FILE_READERS = {
    ".csv": lambda p: pd.read_csv(p, dtype=str),
    ".xlsx": lambda p: pd.read_excel(p, dtype=str),
    ".xls": lambda p: pd.read_excel(p, dtype=str),
}

def _load_table(path: Path) -> pd.DataFrame:
    """Load a CSV/Excel file into a DataFrame with empty strings for NaN."""
    reader = FILE_READERS.get(path.suffix.lower())
    if not reader:
        raise ValueError(f"Unsupported file type: {path.suffix}")
    df = reader(path)
    return df.fillna("")

def append_rows(path: Path, rows: Iterable[Dict[str, Any]], header: List[str]) -> None:
    """Append rows to CSV or Excel, creating file and header if needed."""
    is_excel = path.suffix.lower() in {".xlsx", ".xls"}
    exists = path.exists() and path.stat().st_size > 0
    if is_excel:
        wb = load_workbook(path) if exists else Workbook()
        ws = wb.active
        if not exists:
            ws.append(header)
        for row in rows:
            ws.append([row.get(k, "") for k in header])
        wb.save(path)
    else:
        mode = "a" if exists else "w"
        with open(path, mode, encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=header, extrasaction="ignore")
            if not exists:
                writer.writeheader()
            writer.writerows({k: r.get(k, "") for k in header} for r in rows)

def read_rows(filename: str) -> List[Dict[str, str]]:
    """Read rows from a CSV or Excel file as list of dicts."""
    df = _load_table(Path(filename))
    return df.to_dict(orient="records")

def read_prompt(filename: str) -> str:
    """Read the prompt template from a text file."""
    with open(filename, "r", encoding="utf-8") as f:
        return f.read()

def choose_prompt_file(input_path: Path, user_prompt: Optional[str]) -> Path:
    """Return the prompt file to use, with auto-discovery if none provided."""
    path: Optional[Path] = None
    if user_prompt:
        path = Path(user_prompt).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Prompt file not found: {path}")
        return path

    c1 = input_path.with_suffix(ALT_PROMPT_SUFFIX)
    if c1.exists():
        return c1

    c2 = Path(DEFAULT_PROMPT_FILENAME)
    if c2.exists():
        return c2

    raise FileNotFoundError(
        f"No prompt file supplied and neither '{c1.name}' nor '{c2.name}' exist."
    )

def default_output_file(input_path: Path, user_output: Optional[str]) -> Path:
    """Determine the default output file path."""
    if user_output:
        return Path(user_output)
    suffix = OUTPUT_FILE_SUFFIX
    if input_path.suffix.lower() in {".xlsx", ".xls"} and suffix.endswith(".csv"):
        suffix = suffix[:-4] + ".xlsx"
    return input_path.with_name(f"{input_path.stem}{suffix}")

def collect_existing_ids_and_header(output_file: Path) -> Tuple[Set[str], List[str]]:
    """Return the set of IDs already in output and the existing header (if any)."""
    if not (output_file.exists() and output_file.stat().st_size > 0):
        return set(), []
    df = _load_table(output_file)
    header = list(df.columns)
    if "id" in header:
        ids = set(df["id"].astype(str))
    else:
        ids = {str(i) for i in range(len(df))}
    return ids, header

